# -*- coding: utf-8 -*-
"""
ARINC429 运行时模块 - 固定通用部分
此模块包含所有 ARINC429 解析所需的基础工具函数，
由代码生成器生成的协议解析脚本导入使用。

========================================================
ARINC429 数据字格式 (32位整型):
========================================================
  Bit  1-8  : 标签 (Label), 八进制编码 (位反序)
  Bit  9-10 : SDI (源/目标标识符), 部分字中为预留
  Bit 11-29 : 数据域 (具体定义因数据字类型而异)
  Bit 29    : 符号位 (BNR有符号类型: 0=正, 1=负)
  Bit 30-31 : 状态矩阵 (SSM): 00=故障, 01=无效, 10=测试, 11=正常
  Bit 32    : 奇校验位
"""

import re
from datetime import datetime


# ============================================================
# 基础工具函数
# ============================================================

def reverse_bits_8(byte_val):
    """反转8位二进制的位序 (ARINC429 Label位反序规则)
    
    ARINC429中Label字段(bit1-bit8)的位序与实际Label值相反:
    bit1是Label最高位, bit8是Label最低位.
    
    Args:
        byte_val: 从32位字中直接提取的bits 1-8 (整数0-255)
    Returns:
        位反转后的Label值 (十进制), 用oct()可转为八进制
    """
    result = 0
    for i in range(8):
        if byte_val & (1 << i):
            result |= (1 << (7 - i))
    return result


def extract_label(word):
    """从32位ARINC429字中提取Label值
    
    Args:
        word: 32位整数 (bit1在最低位, bit32在最高位)
    Returns:
        label_decimal: Label的十进制值
        label_octal_str: Label的八进制字符串 (如 '115')
    """
    raw = word & 0xFF  # 提取bits 1-8
    label_val = reverse_bits_8(raw)
    label_oct_str = oct(label_val)[2:]  # 去掉 '0o' 前缀
    return label_val, label_oct_str


def extract_bit(word, bit_num):
    """提取ARINC429字中指定位的值
    
    Args:
        word: 32位整数
        bit_num: ARINC429位号 (1-32, bit1在最低位)
    Returns:
        0 或 1
    """
    return (word >> (bit_num - 1)) & 1


def extract_bits(word, start_bit, end_bit):
    """提取ARINC429字中指定范围的位值 (含两端)
    
    Args:
        word: 32位整数
        start_bit: 起始位号 (1-based, 含)
        end_bit: 结束位号 (1-based, 含)
    Returns:
        提取的无符号整数值 (start_bit为LSB)
    """
    num_bits = end_bit - start_bit + 1
    mask = (1 << num_bits) - 1
    return (word >> (start_bit - 1)) & mask


def check_odd_parity(word):
    """检查32位字的奇校验
    
    Returns:
        True: 校验通过 (1的个数为奇数)
        False: 校验失败
    """
    count = bin(word & 0xFFFFFFFF).count('1')
    return count % 2 == 1


def decode_ssm(ssm_val):
    """解码状态矩阵 (SSM, Bits 30-31)
    
    按协议文档定义:
      00 → 故障
      01 → 无效
      10 → 测试
      11 → 正常
    
    Args:
        ssm_val: 0-3 (bit31为高位, bit30为低位)
    Returns:
        状态描述字符串, 格式: "XX-含义"
    """
    ssm_map = {
        0b00: '00-故障',
        0b01: '01-无效',
        0b10: '10-测试',
        0b11: '11-正常'
    }
    return ssm_map.get(ssm_val, f'{ssm_val:02b}-未知')


def decode_bnr_signed(word, data_start, data_end, sign_bit, resolution):
    """解码BNR有符号数据
    
    使用二进制补码:
    - 符号位=0: 正数, 值 = data × resolution
    - 符号位=1: 负数, 将(符号位+数据位)视为补码整数
    
    Args:
        word: 32位ARINC429字
        data_start: 数据域起始位 (LSB)
        data_end: 数据域结束位 (MSB)
        sign_bit: 符号位位号
        resolution: 分辨率 (物理量/bit)
    Returns:
        (raw_data, sign, physical_value)
    """
    data_raw = extract_bits(word, data_start, data_end)
    sign = extract_bit(word, sign_bit)
    num_data_bits = data_end - data_start + 1
    
    # 组合符号位和数据位, 形成补码整数
    combined = (sign << num_data_bits) | data_raw
    total_bits = num_data_bits + 1
    
    if sign:
        # 负数: 二进制补码转换
        signed_val = combined - (1 << total_bits)
    else:
        signed_val = combined
    
    physical_val = signed_val * resolution
    return data_raw, sign, physical_val


def decode_bnr_unsigned(word, data_start, data_end, resolution):
    """解码BNR无符号数据
    
    Args:
        word: 32位ARINC429字
        data_start: 数据域起始位 (LSB)
        data_end: 数据域结束位 (MSB)
        resolution: 分辨率 (物理量/bit)
    Returns:
        (raw_data, physical_value)
    """
    data_raw = extract_bits(word, data_start, data_end)
    physical_val = data_raw * resolution
    return data_raw, physical_val


def interpret_discrete_desc(desc_str, bit_val):
    """根据离散位描述字符串和当前位值, 返回字段名和含义解释
    
    描述格式示例: "调零指令: 0=调零, 1=无效"
                 "转弯断开开关: 1=转弯断开, 0=无效 (驾驶舱显控开关)"
    
    Args:
        desc_str: 协议中对该位的描述 (如 "调零指令: 0=调零, 1=无效")
        bit_val: 当前位的值 (0 或 1)
    Returns:
        (field_name, interpretation): 字段名称和当前值的含义
    """
    # 提取字段名 (冒号前面的部分)
    if ':' in desc_str:
        field_name = desc_str.split(':')[0].strip()
        rest = desc_str.split(':', 1)[1].strip()
    elif '：' in desc_str:
        field_name = desc_str.split('：')[0].strip()
        rest = desc_str.split('：', 1)[1].strip()
    else:
        field_name = desc_str
        rest = ''
    
    # 尝试从描述中解析 "0=xxx, 1=yyy" 格式
    interpretation = ''
    # 匹配 "0=xxx" 和 "1=yyy" 模式
    matches = re.findall(r'(\d)\s*[=＝]\s*([^,，;；()（）]+)', rest)
    val_map = {}
    for m_val, m_desc in matches:
        val_map[int(m_val)] = m_desc.strip()
    
    if bit_val in val_map:
        interpretation = val_map[bit_val]
    else:
        interpretation = str(bit_val)
    
    return field_name, interpretation


# ============================================================
# 文件读取函数
# ============================================================

def parse_hex_input(hex_str):
    """解析用户输入的十六进制字符串
    
    支持格式:
    - "67FF00B2"       (纯hex, 8位)
    - "0x67FF00B2"     (带0x前缀)
    - "B2 00 FF 67"    (4个空格分隔的字节, 小端序: byte0=bit1-8在前)
    
    Args:
        hex_str: 十六进制字符串
    Returns:
        32位整数
    """
    hex_str = hex_str.strip()
    if hex_str.startswith(('0x', '0X')):
        hex_str = hex_str[2:]
    
    # 检查是否是空格分隔的字节格式 (如 "B2 00 FF 67")
    parts = hex_str.split()
    if len(parts) == 4 and all(len(p) == 2 for p in parts):
        # 4个字节, 小端序: byte0=bits1-8(Label), byte3=bits25-32(SSM+Parity)
        b = bytes([int(p, 16) for p in parts])
        return int.from_bytes(b, byteorder='little')
    
    return int(hex_str, 16)


def load_raw_byte_file(filepath):
    """从原始字节文件中读取ARINC429数据字
    
    支持格式: 空格分隔的十六进制字节流, 每4个字节为一个32位ARINC429字
    例: "B2 00 FF 67 92 00 FE 6B ..."
    
    字节序: 小端 (byte0=bits1-8 即Label, byte3=bits25-32 即SSM+Parity)
    
    Args:
        filepath: 原始数据文件路径
    Returns:
        32位整数列表
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    
    # 提取所有合法的十六进制token
    tokens = []
    for token in content.split():
        token = token.strip()
        if len(token) == 2:
            try:
                int(token, 16)
                tokens.append(token)
            except ValueError:
                continue
    
    # 每4个字节组成一个32位ARINC429字 (小端序)
    words = []
    for i in range(0, len(tokens) - 3, 4):
        try:
            b = bytes([int(tokens[i], 16), int(tokens[i+1], 16),
                        int(tokens[i+2], 16), int(tokens[i+3], 16)])
            word = int.from_bytes(b, byteorder='little')
            words.append(word)
        except (ValueError, IndexError):
            continue
    
    return words


# ============================================================
# Excel 输出函数
# ============================================================

def create_excel_workbook():
    """创建一个新的Excel工作簿，返回(workbook, worksheet, headers)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("需要安装openpyxl: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = '解析结果'
    
    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写表头
    headers = [
        '序号', '原始数据(HEX)', '原始4字节(小端)', '32位二进制',
        'Label(八进制)', '信号名称', '方向',
        '数据类型', '原始数据值', '符号', '物理值', '单位',
        'SSM状态', '奇校验', '备注'
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    return wb, ws, headers


def write_excel_row(ws, row_num, result, word, label_lookup):
    """将一条解析结果写入Excel行
    
    Args:
        ws: worksheet对象
        row_num: 行号 (1-based, 数据从第2行开始)
        result: parse_arinc429_word()返回的结果字典
        word: 原始32位整数
        label_lookup: Label定义查找表
    """
    from openpyxl.styles import Font, Border, Side
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    col = 1
    idx = row_num - 1  # 序号
    
    # 序号
    ws.cell(row=row_num, column=col, value=idx).border = thin_border; col += 1
    # 原始HEX (32位整数)
    ws.cell(row=row_num, column=col, value=result['raw_hex']).border = thin_border; col += 1
    # 原始4字节(小端序): byte0 byte1 byte2 byte3
    raw_bytes = (word & 0xFFFFFFFF).to_bytes(4, byteorder='little')
    raw_4byte_str = ' '.join(f'{b:02X}' for b in raw_bytes)
    ws.cell(row=row_num, column=col, value=raw_4byte_str).border = thin_border; col += 1
    # 32位二进制
    ws.cell(row=row_num, column=col, value=result['raw_bin']).border = thin_border; col += 1
    # Label
    ws.cell(row=row_num, column=col, value=result['label_oct']).border = thin_border; col += 1
    # 信号名称
    ws.cell(row=row_num, column=col, value=result['name']).border = thin_border; col += 1
    # 方向
    ws.cell(row=row_num, column=col, value=result.get('direction', '')).border = thin_border; col += 1
    
    # 数据类型 & 值 (columns 8-12)
    if result['known']:
        word_def = label_lookup.get(result['label_dec'], {})
        
        # 确定数据类型显示
        has_bnr = bool(result.get('bnr_fields'))
        has_discrete = bool(result.get('discrete_bits'))
        has_special = bool(result.get('special_fields'))
        
        if has_bnr and not has_discrete and not has_special:
            # 纯BNR类型
            bf = result['bnr_fields'][0] if result['bnr_fields'] else {}
            if bf.get('sign') is not None:
                data_type = 'BNR_SIGNED'
            else:
                data_type = 'BNR_UNSIGNED'
        elif has_discrete or has_special:
            data_type = 'DISCRETE'
        else:
            data_type = word_def.get('data_type', '未知')
        
        ws.cell(row=row_num, column=col, value=data_type).border = thin_border; col += 1
        
        # 处理 BNR 数值字段
        if has_bnr:
            bnr_parts = []
            phys_parts = []
            unit_str = ''
            for bf in result['bnr_fields']:
                bnr_parts.append(f'{bf["name"]}={bf["data_raw"]}')
                sign_str = f'({bf["sign_desc"]})' if bf.get('sign_desc') else ''
                phys_parts.append(f'{bf["name"]}: {bf["physical_value"]:.6f}{sign_str}')
                if bf.get('unit'):
                    unit_str = bf['unit']
            
            ws.cell(row=row_num, column=col, value='; '.join(bnr_parts)).border = thin_border; col += 1
            # 符号列
            if result['bnr_fields'] and result['bnr_fields'][0].get('sign_desc'):
                ws.cell(row=row_num, column=col, value=result['bnr_fields'][0]['sign_desc']).border = thin_border
            else:
                ws.cell(row=row_num, column=col, value='').border = thin_border
            col += 1
            # 物理值
            if len(result['bnr_fields']) == 1:
                ws.cell(row=row_num, column=col, value=round(result['bnr_fields'][0]['physical_value'], 6)).border = thin_border
            else:
                ws.cell(row=row_num, column=col, value='; '.join(phys_parts)).border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value=unit_str).border = thin_border; col += 1
            
        elif has_discrete or has_special:
            # 离散量: 原始数据值 - 显示各位的定义和值
            raw_parts = []
            interp_parts = []
            for db in result.get('discrete_bits', []):
                field_name, interp = interpret_discrete_desc(db['description'], db['value'])
                raw_parts.append(f'Bit{db["bit"]}={db["value"]}({field_name})')
                interp_parts.append(f'{field_name}: {interp}')
            for sf in result.get('special_fields', []):
                raw_val = sf['raw_value']
                raw_parts.append(f'{sf["name"]}({sf["bits"]})={raw_val}')
                interp_parts.append(f'{sf["name"]}: {sf["description"]}')
            ws.cell(row=row_num, column=col, value='; '.join(raw_parts)).border = thin_border; col += 1
            ws.cell(row=row_num, column=col, value='').border = thin_border; col += 1
            ws.cell(row=row_num, column=col, value='; '.join(interp_parts)).border = thin_border; col += 1
            ws.cell(row=row_num, column=col, value='').border = thin_border; col += 1
        else:
            # 兼容旧格式
            if 'physical_value' in result:
                ws.cell(row=row_num, column=col, value=result.get('data_raw', '')).border = thin_border; col += 1
                sign_str = result.get('sign_desc', '')
                ws.cell(row=row_num, column=col, value=sign_str).border = thin_border; col += 1
                ws.cell(row=row_num, column=col, value=round(result['physical_value'], 6)).border = thin_border; col += 1
                ws.cell(row=row_num, column=col, value=result.get('unit', '')).border = thin_border; col += 1
            else:
                for _ in range(4):
                    ws.cell(row=row_num, column=col, value='').border = thin_border; col += 1
    else:
        ws.cell(row=row_num, column=col, value='未知').border = thin_border; col += 1
        for _ in range(4):
            ws.cell(row=row_num, column=col, value='').border = thin_border; col += 1
    
    # SSM状态
    ws.cell(row=row_num, column=col, value=result['ssm_desc']).border = thin_border; col += 1
    # 校验
    parity_str = '通过' if result['parity_ok'] else '失败'
    cell = ws.cell(row=row_num, column=col, value=parity_str)
    cell.border = thin_border
    if not result['parity_ok']:
        cell.font = Font(color='FF0000', bold=True)
    col += 1
    # 备注
    ws.cell(row=row_num, column=col, value=result.get('notes', '')).border = thin_border


def finalize_excel(wb, ws, output_path):
    """完成Excel并保存
    
    Args:
        wb: workbook对象
        ws: worksheet对象
        output_path: 输出文件路径
    """
    from openpyxl.utils import get_column_letter
    
    # 调整列宽
    col_widths = [6, 14, 16, 38, 12, 25, 14, 14, 45, 6, 45, 8, 25, 8, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    wb.save(output_path)
    return output_path
