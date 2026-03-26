# -*- coding: utf-8 -*-
"""
批量解析IRS惯导原始数据文件
读取IRS2_1.txt，按EB90包头划分数据包，逐包解析并输出到Excel
"""

import struct
import os
import sys
import time
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ========== CRC16-XMODEM ==========
def crc16(data: bytes) -> int:
    """CRC-16/XMODEM: poly=0x1021, init=0x0000"""
    crc = 0x0000
    for byte in data:
        crc ^= byte << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = (crc << 1) ^ 0x1021
            else:
                crc <<= 1
    return crc & 0xFFFF


# ========== 快速解析单个80字节包，返回一行数据列表 ==========
MANUFACTURER_MAP = {0: '航天时代', 1: '陕西华燕', 2: '中科导控'}
DIFF_MAP = {0x04: '无效解', 0x08: '单点定位解', 0x0C: '伪距差分', 0x15: '固定解', 0x0D: '浮点解'}
WORK_MODE_MAP = {0: '准备', 1: '对准', 2: '导航'}
NAV_MODE_MAP = {0: '无导航', 1: '纯惯性', 2: '组合导航'}
ALIGN_STATUS_MAP = {0: '未对准', 1: '对准中', 2: '对准失败', 3: '对准成功'}
ALIGN_MODE_MAP = {0: '静基座', 1: '动基座'}
VALID_MAP = {0x01: '无效', 0x03: '有效'}


def parse_packet_fast(raw: bytes):
    """
    快速解析80字节数据包，返回一个列表（用于Excel的一行）
    """
    if len(raw) < 80:
        return None

    # 包头验证
    header_ok = (raw[0] == 0xEB and raw[1] == 0x90)

    # 基本信息
    class_val = raw[2]
    id_val = raw[3]
    mfr_code = id_val & 0x03
    manufacturer = MANUFACTURER_MAP.get(mfr_code, f'未知({mfr_code})')
    frame_len = raw[4]
    frame_count = raw[5]

    # 惯导数据
    heading = struct.unpack_from('<H', raw, 6)[0] * 0.01
    pitch = struct.unpack_from('<h', raw, 8)[0] * 0.01
    roll = struct.unpack_from('<h', raw, 10)[0] * 0.01
    east_vel = struct.unpack_from('<h', raw, 12)[0] * 0.01
    north_vel = struct.unpack_from('<h', raw, 14)[0] * 0.01
    up_vel = struct.unpack_from('<h', raw, 16)[0] * 0.01
    latitude = struct.unpack_from('<i', raw, 18)[0] * 0.0000001
    longitude = struct.unpack_from('<i', raw, 22)[0] * 0.0000001
    altitude = struct.unpack_from('<i', raw, 26)[0] * 0.01
    gyro_x = struct.unpack_from('<h', raw, 30)[0] * 0.01
    gyro_y = struct.unpack_from('<h', raw, 32)[0] * 0.01
    gyro_z = struct.unpack_from('<h', raw, 34)[0] * 0.01
    acc_x = struct.unpack_from('<h', raw, 36)[0] * 0.01
    acc_y = struct.unpack_from('<h', raw, 38)[0] * 0.01
    acc_z = struct.unpack_from('<h', raw, 40)[0] * 0.01

    # 工作状态字1
    byte42 = raw[42]
    byte43 = raw[43]
    work_mode = WORK_MODE_MAP.get(byte42 & 0x03, '未知')
    nav_mode = NAV_MODE_MAP.get((byte42 >> 3) & 0x03, '未知')
    install_comp = '已补偿' if (byte42 >> 5) & 0x01 else '未补偿'
    gnss_src = (byte42 >> 6) & 0x03
    align_status = ALIGN_STATUS_MAP.get(byte43 & 0x03, '未知')
    align_mode = ALIGN_MODE_MAP.get((byte43 >> 2) & 0x01, '未知')

    # 故障字
    byte44 = raw[44]
    fault1_self = '故障' if (byte44 & 0x01) else '正常'
    fault1_init = '故障' if (byte44 & 0x02) else '正常'

    byte45 = raw[45]
    fault2_str = '/'.join([
        '故障' if (byte45 & (1 << i)) else '正常' for i in range(6)
    ])  # X陀/Y陀/Z陀/X加/Y加/Z加

    fault3 = struct.unpack_from('<H', raw, 46)[0]

    # 转发卫导数据
    hdop_rtk1 = (struct.unpack_from('<I', raw, 48)[0] & 0xFFFF) * 0.03125
    hdop_rtk2 = (struct.unpack_from('<I', raw, 52)[0] & 0xFFFF) * 0.03125
    vdop_rtk1 = (struct.unpack_from('<I', raw, 56)[0] & 0xFFFF) * 0.03125
    vdop_rtk2 = (struct.unpack_from('<I', raw, 60)[0] & 0xFFFF) * 0.03125
    sat_rtk1 = raw[64] & 0x1F
    sat_rtk2 = raw[65] & 0x1F

    diff_rtk1_code = struct.unpack_from('<H', raw, 66)[0] & 0x1F
    diff_rtk1 = DIFF_MAP.get(diff_rtk1_code, f'未知(0x{diff_rtk1_code:02X})')
    diff_rtk2_code = struct.unpack_from('<H', raw, 68)[0] & 0x1F
    diff_rtk2 = DIFF_MAP.get(diff_rtk2_code, f'未知(0x{diff_rtk2_code:02X})')

    gnss_v1 = raw[70]
    gnss_status1 = VALID_MAP.get(gnss_v1 & 0x03, '未知')
    dop_status1 = VALID_MAP.get((gnss_v1 >> 2) & 0x03, '未知')
    gnss_v2 = raw[71]
    gnss_status2 = VALID_MAP.get(gnss_v2 & 0x03, '未知')
    dop_status2 = VALID_MAP.get((gnss_v2 >> 2) & 0x03, '未知')

    # 版本号
    sw_raw = struct.unpack_from('<H', raw, 72)[0]
    sw_minor = sw_raw & 0x3F
    sw_major = (sw_raw >> 6) & 0x0F
    sw_mfr = (sw_raw >> 10) & 0x03
    sw_ver = f'{MANUFACTURER_MAP.get(sw_mfr, "?")}-SW-V{sw_major:02d}.{sw_minor:03d}'

    hw_raw = struct.unpack_from('<H', raw, 74)[0]
    hw_minor = hw_raw & 0x3F
    hw_major = (hw_raw >> 6) & 0x0F
    hw_mfr = (hw_raw >> 10) & 0x03
    hw_ver = f'{MANUFACTURER_MAP.get(hw_mfr, "?")}-HW-V{hw_major:02d}.{hw_minor:03d}'

    reserved = struct.unpack_from('<H', raw, 76)[0]

    # CRC
    crc_recv = struct.unpack_from('<H', raw, 78)[0]
    crc_calc = crc16(raw[2:78])
    crc_ok = '通过' if crc_recv == crc_calc else '失败'

    # 原始hex
    raw_hex = raw.hex().upper()

    return [
        header_ok, manufacturer, frame_len, frame_count,
        heading, pitch, roll,
        east_vel, north_vel, up_vel,
        latitude, longitude, altitude,
        gyro_x, gyro_y, gyro_z,
        acc_x, acc_y, acc_z,
        work_mode, nav_mode, install_comp, gnss_src,
        align_status, align_mode,
        fault1_self, fault1_init, fault2_str, fault3,
        hdop_rtk1, hdop_rtk2, vdop_rtk1, vdop_rtk2,
        sat_rtk1, sat_rtk2,
        diff_rtk1, diff_rtk2,
        gnss_status1, dop_status1, gnss_status2, dop_status2,
        sw_ver, hw_ver, reserved,
        f'0x{crc_recv:04X}', f'0x{crc_calc:04X}', crc_ok,
        raw_hex
    ]


# Excel表头
HEADERS = [
    '包序号', '包头正确', '厂家', '帧长度', '帧计数',
    '航向(°)', '俯仰(°)', '滚动(°)',
    '东速(m/s)', '北速(m/s)', '天速(m/s)',
    '纬度(°)', '经度(°)', '高度(m)',
    'X角速度(°/s)', 'Y角速度(°/s)', 'Z角速度(°/s)',
    'X加速度(m/s²)', 'Y加速度(m/s²)', 'Z加速度(m/s²)',
    '工作方式', '导航模式', '安装误差补偿', '卫导数据源',
    '对准状态', '对准方式',
    '故障1-周期自检', '故障1-开机初始化',
    '故障2(X陀/Y陀/Z陀/X加/Y加/Z加)', '故障字3(raw)',
    '水平精度RTK1', '水平精度RTK2', '垂直精度RTK1', '垂直精度RTK2',
    '星数RTK1', '星数RTK2',
    '差分RTK1', '差分RTK2',
    '卫导状态RTK1', 'DOP状态RTK1', '卫导状态RTK2', 'DOP状态RTK2',
    '软件版本', '硬件版本', '预留',
    'CRC接收', 'CRC计算', 'CRC校验',
    '原始数据(hex)'
]

MAX_ROWS_PER_SHEET = 1000000  # 留余量，Excel最大1048576行


def main():
    import io
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    # 找文件
    desktop = r'C:\Users\wangz\Desktop'
    target = None
    for f in os.listdir(desktop):
        full = os.path.join(desktop, f)
        if os.path.isdir(full):
            try:
                for inner_f in os.listdir(full):
                    if 'IRS' in inner_f and inner_f.endswith('.txt'):
                        target = os.path.join(full, inner_f)
            except:
                pass

    if not target:
        print("未找到IRS2_1.txt文件！")
        return

    file_size = os.path.getsize(target)
    print(f"文件: {target}")
    print(f"大小: {file_size / 1024 / 1024:.2f} MB")

    # 读取并解析
    print("\n正在读取文件...")
    start_time = time.time()

    # 分块读取文本，提取所有token
    token_buffer = []
    chunk_size = 50 * 1024 * 1024  # 50MB per chunk

    with open(target, 'r', encoding='utf-8', errors='replace') as f:
        leftover = ''
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            text = leftover + chunk
            # 找到最后一个空格以避免切割token
            last_space = text.rfind(' ')
            if last_space == -1 or not chunk:
                process_text = text
                leftover = ''
            else:
                process_text = text[:last_space]
                leftover = text[last_space + 1:]

            tokens = process_text.split()
            token_buffer.extend(tokens)

            elapsed = time.time() - start_time
            print(f"\r  已读取 {f.tell() / 1024 / 1024:.0f} MB / {file_size / 1024 / 1024:.0f} MB  "
                  f"累计tokens: {len(token_buffer):,}  耗时: {elapsed:.1f}s", end='', flush=True)

        # Process remaining leftover
        if leftover.strip():
            token_buffer.extend(leftover.split())

    print(f"\n读取完成！总tokens: {len(token_buffer):,}")

    # 智能查找数据包：用CRC校验区分真假EB90包头
    print("正在查找并验证数据包（CRC校验区分真假包头）...")
    parsed_rows = []
    skipped_false_eb90 = 0
    errors = 0
    
    i = 0
    total_tokens = len(token_buffer)
    while i < total_tokens - 79:
        # 寻找EB 90
        if token_buffer[i].upper() != 'EB' or token_buffer[i + 1].upper() != '90':
            i += 1
            continue
        
        # 找到EB 90，尝试提取80字节并验证CRC
        try:
            raw = bytes([int(token_buffer[i + j], 16) for j in range(80)])
        except (ValueError, IndexError):
            errors += 1
            i += 1
            continue
        
        # CRC校验：用CRC判断这是否是真正的包头
        crc_recv = struct.unpack_from('<H', raw, 78)[0]
        crc_calc = crc16(raw[2:78])
        
        if crc_recv == crc_calc:
            # CRC通过 → 真正的数据包
            row = parse_packet_fast(raw)
            if row is not None:
                parsed_rows.append([len(parsed_rows) + 1] + row)
            i += 80  # 跳过整个包，从下一个包开始找
        else:
            # CRC不通过 → 数据内容中的假EB90，跳过
            skipped_false_eb90 += 1
            i += 1  # 只前进1步，继续寻找真正的包头
        
        if len(parsed_rows) % 100000 == 0 and len(parsed_rows) > 0:
            elapsed = time.time() - start_time
            print(f"\r  已找到 {len(parsed_rows):,} 个有效包  跳过假包头: {skipped_false_eb90}  "
                  f"进度: {i/total_tokens*100:.1f}%  耗时: {elapsed:.1f}s", end='', flush=True)

    elapsed = time.time() - start_time
    print(f"\n解析完成！")
    print(f"  有效数据包: {len(parsed_rows):,}")
    print(f"  跳过的假EB90: {skipped_false_eb90}")
    print(f"  解析异常: {errors}")
    print(f"  耗时: {elapsed:.1f}s")

    # 释放token_buffer内存
    del token_buffer

    # 写入Excel
    print("\n正在生成Excel文件...")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = os.path.dirname(target)
    output_path = os.path.join(output_dir, f'IRS2_1_解析结果_{timestamp}.xlsx')

    wb = Workbook(write_only=True)

    # 样式（write_only模式不支持后续修改样式，需在写入时设置）
    header_font = Font(name='微软雅黑', bold=True, size=10)
    data_font = Font(name='微软雅黑', size=9)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    sheet_count = 0
    row_in_sheet = 0
    ws = None

    for i, row in enumerate(parsed_rows):
        if i % MAX_ROWS_PER_SHEET == 0:
            sheet_count += 1
            ws = wb.create_sheet(title=f'数据_{sheet_count}')
            # 写表头
            ws.append(HEADERS)
            row_in_sheet = 1

        ws.append(row)
        row_in_sheet += 1

        if i % 200000 == 0 and i > 0:
            elapsed = time.time() - start_time
            print(f"\r  已写入 {i:,} / {len(parsed_rows):,} 行  耗时: {elapsed:.1f}s", end='', flush=True)

    print(f"\n正在保存文件（这可能需要几分钟）...")
    save_start = time.time()
    wb.save(output_path)
    save_time = time.time() - save_start

    total_time = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  处理完成！")
    print(f"  有效数据包: {len(parsed_rows):,}")
    print(f"  跳过假EB90: {skipped_false_eb90}")
    print(f"  CRC全部通过: 是")
    print(f"  Excel sheets: {sheet_count}")
    print(f"  输出文件: {output_path}")
    print(f"  文件保存耗时: {save_time:.1f}s")
    print(f"  总耗时: {total_time:.1f}s")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
