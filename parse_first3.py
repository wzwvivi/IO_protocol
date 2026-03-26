# -*- coding: utf-8 -*-
"""
惯导通讯协议解析 - 只解析前3个数据包
每个包一个sheet，格式与原协议文档一致
"""
import struct
import os
import sys
import io
from datetime import datetime

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ==================== CRC16-XMODEM ====================
def crc16(data: bytes) -> int:
    """CRC-16/XMODEM: poly=0x1021, init=0x0000"""
    crc = 0x0000
    for byte in data:
        crc ^= byte << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = (crc << 1) ^ 0x1021
            else:
                crc <<= 1
    return crc & 0xFFFF


# ==================== 查找文件 ====================
def find_irs_file():
    desktop = r'C:\Users\wangz\Desktop'
    for f in os.listdir(desktop):
        full = os.path.join(desktop, f)
        if os.path.isdir(full):
            try:
                for inner_f in os.listdir(full):
                    if 'IRS' in inner_f and inner_f.endswith('.txt'):
                        return os.path.join(full, inner_f)
            except:
                pass
    return None


# ==================== 提取前N个包 ====================
def extract_packets(filepath, n=3):
    """从文件中提取前n个以EB 90开头的80字节包"""
    # 只读取前面一小部分就够了（前3个包大概需要 3*80*3 = 720 字节文本）
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        text = f.read(50000)  # 50KB足够找到前3个包

    tokens = text.strip().split()
    
    packets = []
    i = 0
    while i < len(tokens) - 1 and len(packets) < n:
        if tokens[i].upper() == 'EB' and tokens[i + 1].upper() == '90':
            if i + 80 <= len(tokens):
                hex_tokens = tokens[i:i + 80]
                try:
                    raw = bytes([int(t, 16) for t in hex_tokens])
                    packets.append(raw)
                except ValueError:
                    pass
            i += 80  # 跳到下一个可能的包
        else:
            i += 1
    
    return packets


# ==================== 完整解析一个包 ====================
MANUFACTURER_MAP = {0: '航天时代', 1: '陕西华燕', 2: '中科导控'}


def parse_packet(raw: bytes) -> dict:
    """完整解析80字节数据包"""
    r = {}
    
    # --- 包头 ---
    r['header1'] = raw[0]
    r['header2'] = raw[1]
    r['header_ok'] = raw[0] == 0xEB and raw[1] == 0x90
    r['class'] = raw[2]
    r['id_raw'] = raw[3]
    r['id_mfr_code'] = raw[3] & 0x03
    r['id_mfr'] = MANUFACTURER_MAP.get(r['id_mfr_code'], f"未知({r['id_mfr_code']})")
    r['frame_len'] = raw[4]
    r['frame_count'] = raw[5]
    
    # --- 惯导数据 ---
    r['heading_raw'] = struct.unpack_from('<H', raw, 6)[0]
    r['heading'] = r['heading_raw'] * 0.01
    
    r['pitch_raw'] = struct.unpack_from('<h', raw, 8)[0]
    r['pitch'] = r['pitch_raw'] * 0.01
    
    r['roll_raw'] = struct.unpack_from('<h', raw, 10)[0]
    r['roll'] = r['roll_raw'] * 0.01
    
    r['east_vel_raw'] = struct.unpack_from('<h', raw, 12)[0]
    r['east_vel'] = r['east_vel_raw'] * 0.01
    
    r['north_vel_raw'] = struct.unpack_from('<h', raw, 14)[0]
    r['north_vel'] = r['north_vel_raw'] * 0.01
    
    r['up_vel_raw'] = struct.unpack_from('<h', raw, 16)[0]
    r['up_vel'] = r['up_vel_raw'] * 0.01
    
    r['lat_raw'] = struct.unpack_from('<i', raw, 18)[0]
    r['latitude'] = r['lat_raw'] * 0.0000001
    r['lat_dir'] = '南纬' if r['lat_raw'] < 0 else '北纬'
    
    r['lon_raw'] = struct.unpack_from('<i', raw, 22)[0]
    r['longitude'] = r['lon_raw'] * 0.0000001
    r['lon_dir'] = '西经' if r['lon_raw'] < 0 else '东经'
    
    r['alt_raw'] = struct.unpack_from('<i', raw, 26)[0]
    r['altitude'] = r['alt_raw'] * 0.01
    
    r['gyro_x_raw'] = struct.unpack_from('<h', raw, 30)[0]
    r['gyro_x'] = r['gyro_x_raw'] * 0.01
    r['gyro_y_raw'] = struct.unpack_from('<h', raw, 32)[0]
    r['gyro_y'] = r['gyro_y_raw'] * 0.01
    r['gyro_z_raw'] = struct.unpack_from('<h', raw, 34)[0]
    r['gyro_z'] = r['gyro_z_raw'] * 0.01
    
    r['acc_x_raw'] = struct.unpack_from('<h', raw, 36)[0]
    r['acc_x'] = r['acc_x_raw'] * 0.01
    r['acc_y_raw'] = struct.unpack_from('<h', raw, 38)[0]
    r['acc_y'] = r['acc_y_raw'] * 0.01
    r['acc_z_raw'] = struct.unpack_from('<h', raw, 40)[0]
    r['acc_z'] = r['acc_z_raw'] * 0.01
    
    # --- 工作状态字1 ---
    r['status1_raw'] = struct.unpack_from('<H', raw, 42)[0]
    b42, b43 = raw[42], raw[43]
    
    wm = b42 & 0x03
    r['work_mode'] = {0: '准备（开机后，对准前）', 1: '对准', 2: '导航'}.get(wm, f'未知({wm})')
    nm = (b42 >> 3) & 0x03
    r['nav_mode'] = {0: '无导航', 1: '纯惯性', 2: '组合导航'}.get(nm, f'未知({nm})')
    ic = (b42 >> 5) & 0x01
    r['install_comp'] = {0: '未补偿', 1: '补偿成功'}.get(ic)
    gs = (b42 >> 6) & 0x03
    r['gnss_source'] = {0: 'NA(非组合导航模式下)', 1: '卫导数据源1', 2: '卫导数据源2'}.get(gs, f'未知({gs})')
    ast = b43 & 0x03
    r['align_status'] = {0: '未（准备）对准', 1: '对准进行中', 2: '对准失败', 3: '对准成功'}.get(ast, f'未知({ast})')
    am = (b43 >> 2) & 0x01
    r['align_mode'] = {0: '静基座对准', 1: '动基座对准'}.get(am)
    aps = (b43 >> 3) & 0x03
    r['align_pos_src'] = {0: '无位置数据', 1: '卫星导航接收机数据', 2: '飞管发送经纬高数据'}.get(aps, f'未知({aps})')
    
    # --- 故障字1 byte44 ---
    r['fault1_raw'] = raw[44]
    r['fault1_selftest'] = '故障' if (raw[44] & 0x01) else '正常'
    r['fault1_init'] = '故障' if (raw[44] & 0x02) else '正常'
    
    # --- 故障字2 byte45 ---
    r['fault2_raw'] = raw[45]
    b45 = raw[45]
    r['f2_x_gyro'] = '故障' if (b45 & 0x01) else '正常'
    r['f2_y_gyro'] = '故障' if (b45 & 0x02) else '正常'
    r['f2_z_gyro'] = '故障' if (b45 & 0x04) else '正常'
    r['f2_x_acc'] = '故障' if (b45 & 0x08) else '正常'
    r['f2_y_acc'] = '故障' if (b45 & 0x10) else '正常'
    r['f2_z_acc'] = '故障' if (b45 & 0x20) else '正常'
    
    # --- 故障字3 bytes46-47 ---
    r['fault3_raw'] = struct.unpack_from('<H', raw, 46)[0]
    b47, b46 = raw[47], raw[46]
    r['f3_attitude'] = '故障' if (b47 & 0x01) else '正常'
    r['f3_heading'] = '故障' if (b47 & 0x02) else '正常'
    r['f3_latlon'] = '故障' if (b47 & 0x04) else '正常'
    r['f3_alt'] = '故障' if (b47 & 0x08) else '正常'
    r['f3_vert_vel'] = '故障' if (b47 & 0x10) else '正常'
    r['f3_east_vel'] = '故障' if (b47 & 0x20) else '正常'
    r['f3_north_vel'] = '故障' if (b47 & 0x40) else '正常'
    r['f3_x_gyro'] = '故障' if (b47 & 0x80) else '正常'
    r['f3_y_gyro'] = '故障' if (b46 & 0x01) else '正常'
    r['f3_z_gyro'] = '故障' if (b46 & 0x02) else '正常'
    r['f3_x_acc'] = '故障' if (b46 & 0x04) else '正常'
    r['f3_y_acc'] = '故障' if (b46 & 0x08) else '正常'
    r['f3_z_acc'] = '故障' if (b46 & 0x10) else '正常'
    
    # --- 卫导数据 ---
    r['hdop_rtk1'] = (struct.unpack_from('<I', raw, 48)[0] & 0xFFFF) * 0.03125
    r['hdop_rtk2'] = (struct.unpack_from('<I', raw, 52)[0] & 0xFFFF) * 0.03125
    r['vdop_rtk1'] = (struct.unpack_from('<I', raw, 56)[0] & 0xFFFF) * 0.03125
    r['vdop_rtk2'] = (struct.unpack_from('<I', raw, 60)[0] & 0xFFFF) * 0.03125
    r['sat_rtk1'] = raw[64] & 0x1F
    r['sat_rtk2'] = raw[65] & 0x1F
    
    diff_map = {0x04: '无效解', 0x08: '单点定位解', 0x0C: '伪距差分', 0x15: '固定解', 0x0D: '浮点解'}
    dc1 = struct.unpack_from('<H', raw, 66)[0] & 0x1F
    r['diff_rtk1'] = diff_map.get(dc1, f'未知(0x{dc1:02X})')
    dc2 = struct.unpack_from('<H', raw, 68)[0] & 0x1F
    r['diff_rtk2'] = diff_map.get(dc2, f'未知(0x{dc2:02X})')
    
    valid_map = {0x01: '无效', 0x03: '有效'}
    gv1 = raw[70]
    r['gnss_status_rtk1'] = valid_map.get(gv1 & 0x03, f'未知({gv1 & 0x03:02b})')
    r['dop_status_rtk1'] = valid_map.get((gv1 >> 2) & 0x03, f'未知({(gv1>>2) & 0x03:02b})')
    gv2 = raw[71]
    r['gnss_status_rtk2'] = valid_map.get(gv2 & 0x03, f'未知({gv2 & 0x03:02b})')
    r['dop_status_rtk2'] = valid_map.get((gv2 >> 2) & 0x03, f'未知({(gv2>>2) & 0x03:02b})')
    
    # --- 版本号 ---
    sw = struct.unpack_from('<H', raw, 72)[0]
    r['sw_minor'] = sw & 0x3F
    r['sw_major'] = (sw >> 6) & 0x0F
    r['sw_mfr'] = (sw >> 10) & 0x03
    r['sw_version'] = f"{MANUFACTURER_MAP.get(r['sw_mfr'], '?')}-SW-V{r['sw_major']:02d}.0{r['sw_minor']:02d}"
    
    hw = struct.unpack_from('<H', raw, 74)[0]
    r['hw_minor'] = hw & 0x3F
    r['hw_major'] = (hw >> 6) & 0x0F
    r['hw_mfr'] = (hw >> 10) & 0x03
    r['hw_version'] = f"{MANUFACTURER_MAP.get(r['hw_mfr'], '?')}-HW-V{r['hw_major']:02d}.0{r['hw_minor']:02d}"
    
    r['reserved'] = struct.unpack_from('<H', raw, 76)[0]
    
    # --- CRC ---
    r['crc_recv'] = struct.unpack_from('<H', raw, 78)[0]
    r['crc_calc'] = crc16(raw[2:78])
    r['crc_ok'] = r['crc_recv'] == r['crc_calc']
    
    r['raw_bytes'] = raw
    return r


# ==================== 生成单个sheet ====================
def write_sheet(ws, pkt_idx, parsed):
    """
    在ws中按协议文档格式写入一个包的解析结果
    左半部分是协议定义，右半部分(H、I列)是解析值
    """
    raw = parsed['raw_bytes']
    
    # 样式
    hdr_font = Font(name='微软雅黑', bold=True, size=10)
    data_font = Font(name='微软雅黑', size=9)
    val_font = Font(name='微软雅黑', size=10, bold=True, color='0000FF')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_font_w = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
    sec_fill = PatternFill('solid', fgColor='D6E4F0')
    val_fill = PatternFill('solid', fgColor='FFF2CC')
    
    # 列宽
    for col, w in {'A': 14, 'B': 8, 'C': 24, 'D': 18, 'E': 7, 'F': 40, 'G': 48, 'H': 18, 'I': 30}.items():
        ws.column_dimensions[col].width = w
    
    # ---------- 写表头 ----------
    row = 1
    headers = ['分类', '字节', '数据', '数据类型', '字节数', '说明', '备注', '原始Hex', '解析值']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font_w
        c.fill = hdr_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 25
    
    # ---------- 辅助写行函数 ----------
    def w_row(r, cat, byte_s, name, dtype, nbytes, desc, note, hex_val, result):
        """写一行"""
        vals = [cat, byte_s, name, dtype, nbytes, desc, note, hex_val, result]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = data_font
            c.border = border
            if ci == 1:
                c.fill = sec_fill
                c.alignment = center
            elif ci in (2, 4, 5):
                c.alignment = center
            elif ci >= 8:
                c.fill = val_fill
                c.font = val_font if ci == 9 else data_font
                c.alignment = left
            else:
                c.alignment = left
    
    def hex_of(start, end):
        """获取字节范围的hex字符串"""
        return ' '.join(f'{b:02X}' for b in raw[start:end+1])
    
    # ---------- 写数据行 ----------
    r = 2  # 当前行号
    
    # ===== 包头区域 =====
    w_row(r, '包头', '0', '包头1', 'unsigned char', 1, '固定0xEB',
          '', f'0x{raw[0]:02X}', '正确' if raw[0]==0xEB else '错误'); r+=1
    
    w_row(r, '', '1', '包头2', 'unsigned char', 1, '固定0x90',
          '', f'0x{raw[1]:02X}', '正确' if raw[1]==0x90 else '错误'); r+=1
    
    w_row(r, '', '2', 'Class', 'unsigned char', 1, '固定0x00',
          '', f'0x{raw[2]:02X}', f'0x{raw[2]:02X}'); r+=1
    
    w_row(r, '', '3', 'ID', 'unsigned char', 1,
          'D0-D1：\n航天时代：00\n陕西华燕：01\n中科导控：10\n其余位置零',
          f"用于区分3台惯导\n{parsed['id_mfr']}",
          f'0x{raw[3]:02X}', f"厂家：{parsed['id_mfr']}"); r+=1
    
    w_row(r, '', '4', '帧长度', 'unsigned char', 1, '80',
          '', f'0x{raw[4]:02X}', str(parsed['frame_len'])); r+=1
    
    w_row(r, '', '5', '帧计数', 'unsigned char', 1, '每发一帧自加1',
          '', f'0x{raw[5]:02X}', str(parsed['frame_count'])); r+=1
    
    # 合并A列"包头"
    ws.merge_cells(start_row=2, start_column=1, end_row=7, end_column=1)
    
    # ===== 惯导数据区域 =====
    nav_start = r
    
    # --- 每个导航字段：3行（字节范围+格式、单位范围、LSB） ---
    def write_nav_3row(start_r, byte_str, name, dtype, nbytes, byte_start, byte_end,
                       unit, rng, lsb, value, extra=''):
        signed = 'short' in dtype.lower() or dtype == 'int'
        line1 = f"{byte_start}B(LSB)-{byte_end}B(MSB)，{'最高位为符号位' if signed else '无符号位'}"
        line2 = f"单位为{unit}（{rng}）{'1负0正' if signed else ''}"
        line3 = f"LSB={lsb}"
        hx = hex_of(int(byte_str.split('-')[0]), int(byte_str.split('-')[-1]))
        val_str = f"{value}{extra}"
        
        w_row(start_r, '', byte_str, name, dtype, nbytes, line1, '', hx, val_str)
        w_row(start_r+1, '', '', '', '', '', line2, '', '', '')
        w_row(start_r+2, '', '', '', '', '', line3, '', '', '')
        # 合并 B-E 和 H-I
        for col in [2, 3, 4, 5, 8, 9]:
            ws.merge_cells(start_row=start_r, start_column=col, end_row=start_r+2, end_column=col)
    
    write_nav_3row(r, '6-7', '航向', 'unsigned short int', 2, 6, 7,
                   '度', '0-360°', '0.01°', f"{parsed['heading']:.2f}°"); r+=3
    write_nav_3row(r, '8-9', '俯仰', 'short int', 2, 8, 9,
                   '度', '-90-90°', '0.01°', f"{parsed['pitch']:.2f}°"); r+=3
    write_nav_3row(r, '10-11', '滚动', 'short int', 2, 10, 11,
                   '度', '-180-180°', '0.01°', f"{parsed['roll']:.2f}°"); r+=3
    write_nav_3row(r, '12-13', '东速', 'short int', 2, 12, 13,
                   'm/s', '-327-327', '0.01 m/s', f"{parsed['east_vel']:.2f} m/s"); r+=3
    write_nav_3row(r, '14-15', '北速', 'short int', 2, 14, 15,
                   'm/s', '-327-327', '0.01 m/s', f"{parsed['north_vel']:.2f} m/s"); r+=3
    write_nav_3row(r, '16-17', '天速', 'short int', 2, 16, 17,
                   'm/s', '-327-327', '0.01 m/s', f"{parsed['up_vel']:.2f} m/s"); r+=3
    write_nav_3row(r, '18-21', '纬度', 'int', 4, 18, 21,
                   '度', '±90°', '0.0000001°',
                   f"{parsed['latitude']:.7f}°", f" ({parsed['lat_dir']})"); r+=3
    write_nav_3row(r, '22-25', '经度', 'int', 4, 22, 25,
                   '度', '±180°', '0.0000001°',
                   f"{parsed['longitude']:.7f}°", f" ({parsed['lon_dir']})"); r+=3
    write_nav_3row(r, '26-29', '高度', 'int', 4, 26, 29,
                   '米', '±', '0.01m', f"{parsed['altitude']:.2f} m"); r+=3
    write_nav_3row(r, '30-31', 'X轴角速度(右-前-上)', 'short int', 2, 30, 31,
                   '°/s', '-100-100', '0.01°/s', f"{parsed['gyro_x']:.2f} °/s"); r+=3
    write_nav_3row(r, '32-33', 'Y轴角速度(右-前-上)', 'short int', 2, 32, 33,
                   '°/s', '-100-100', '0.01°/s', f"{parsed['gyro_y']:.2f} °/s"); r+=3
    write_nav_3row(r, '34-35', 'Z轴角速度(右-前-上)', 'short int', 2, 34, 35,
                   '°/s', '-100-100', '0.01°/s', f"{parsed['gyro_z']:.2f} °/s"); r+=3
    write_nav_3row(r, '36-37', 'X轴加速度(右-前-上)', 'short int', 2, 36, 37,
                   'm/s²', '-100-100', '0.01 m/s²', f"{parsed['acc_x']:.2f} m/s²"); r+=3
    write_nav_3row(r, '38-39', 'Y轴加速度(右-前-上)', 'short int', 2, 38, 39,
                   'm/s²', '-100-100', '0.01 m/s²', f"{parsed['acc_y']:.2f} m/s²"); r+=3
    write_nav_3row(r, '40-41', 'Z轴加速度(右-前-上)', 'short int', 2, 40, 41,
                   'm/s²', '-100-100', '0.01 m/s²', f"{parsed['acc_z']:.2f} m/s²"); r+=3
    
    # --- 工作状态字1 (14行) ---
    status_start = r
    desc_lines = [
        '42B：', 'D0 工作方式LSB', 'D1 工作方式MSB',
        'D3 导航模式LSB', 'D4 导航模式MSB', 'D5 安装误差补偿状态',
        'D6 组合导航当前使用卫导数据源LSB', 'D7 组合导航当前使用卫导数据源MSB',
        '43B：', 'D0 对准状态LSB', 'D1 对准状态MSB',
        'D2 对准方式', 'D3 对准位置来源LSB', 'D4 对准位置来源MSB'
    ]
    result_str = (f"工作方式：{parsed['work_mode']}\n"
                  f"导航模式：{parsed['nav_mode']}\n"
                  f"安装误差补偿：{parsed['install_comp']}\n"
                  f"卫导数据源：{parsed['gnss_source']}\n"
                  f"对准状态：{parsed['align_status']}\n"
                  f"对准方式：{parsed['align_mode']}\n"
                  f"对准位置来源：{parsed['align_pos_src']}")
    note_str = ("工作方式：0-准备,1-对准,2-导航\n"
                "导航模式：0-无导航,1-纯惯性,2-组合导航\n"
                "对准状态：0-未对准,1-进行中,2-失败,3-成功")
    
    for di, dl in enumerate(desc_lines):
        if di == 0:
            w_row(r+di, '', '42-43', '工作状态字1', 'unsigned short int', 2,
                  dl, note_str, hex_of(42, 43), result_str)
        else:
            w_row(r+di, '', '', '', '', '', dl, '', '', '')
    for col in [2, 3, 4, 5, 7, 8, 9]:
        ws.merge_cells(start_row=status_start, start_column=col, end_row=status_start+13, end_column=col)
    r += 14
    
    # --- 故障字1 byte44 (3行) ---
    f1_start = r
    f1_descs = ['44B：', 'D0 周期自检测状态：1-故障，0-正常', 'D1 开机初始化状态：1-故障，0-正常']
    f1_result = f"周期自检：{parsed['fault1_selftest']}\n开机初始化：{parsed['fault1_init']}"
    for di, dl in enumerate(f1_descs):
        if di == 0:
            w_row(r+di, '', '44', '故障字1', 'unsigned char', 1, dl, '', f'0x{raw[44]:02X}', f1_result)
        else:
            w_row(r+di, '', '', '', '', '', dl, '', '', '')
    for col in [2, 3, 4, 5, 8, 9]:
        ws.merge_cells(start_row=f1_start, start_column=col, end_row=f1_start+2, end_column=col)
    r += 3
    
    # --- 故障字2 byte45 (6行) ---
    f2_start = r
    f2_descs = ['D0 X陀螺状态', 'D1 Y陀螺状态', 'D2 Z陀螺状态',
                'D3 X加表状态', 'D4 Y加表状态', 'D5 Z加表状态']
    f2_result = (f"X陀螺：{parsed['f2_x_gyro']}\nY陀螺：{parsed['f2_y_gyro']}\nZ陀螺：{parsed['f2_z_gyro']}\n"
                 f"X加表：{parsed['f2_x_acc']}\nY加表：{parsed['f2_y_acc']}\nZ加表：{parsed['f2_z_acc']}")
    for di, dl in enumerate(f2_descs):
        if di == 0:
            w_row(r+di, '', '45', '故障字2', 'unsigned char', 1, dl, '', f'0x{raw[45]:02X}', f2_result)
        else:
            w_row(r+di, '', '', '', '', '', dl, '', '', '')
    for col in [2, 3, 4, 5, 8, 9]:
        ws.merge_cells(start_row=f2_start, start_column=col, end_row=f2_start+5, end_column=col)
    r += 6
    
    # --- 故障字3 bytes46-47 (15行) ---
    f3_start = r
    f3_descs = ['47B:', 'D0 姿态有效性', 'D1 航向角有效性', 'D2 经纬度有效性',
                'D3 高度有效性', 'D4 升降速度有效性', 'D5 东向速度有效性',
                'D6 北向速度有效性', 'D7 x轴角速度有效性',
                '48B:', 'D0 y轴角速度有效性', 'D1 z轴角速度有效性',
                'D2 x轴加速度有效性', 'D3 y轴加速度有效性', 'D4 z轴加速度有效性']
    f3_result = (f"姿态：{parsed['f3_attitude']}\n航向角：{parsed['f3_heading']}\n"
                 f"经纬度：{parsed['f3_latlon']}\n高度：{parsed['f3_alt']}\n"
                 f"升降速度：{parsed['f3_vert_vel']}\n东向速度：{parsed['f3_east_vel']}\n"
                 f"北向速度：{parsed['f3_north_vel']}\nX角速度：{parsed['f3_x_gyro']}\n"
                 f"Y角速度：{parsed['f3_y_gyro']}\nZ角速度：{parsed['f3_z_gyro']}\n"
                 f"X加速度：{parsed['f3_x_acc']}\nY加速度：{parsed['f3_y_acc']}\n"
                 f"Z加速度：{parsed['f3_z_acc']}")
    for di, dl in enumerate(f3_descs):
        if di == 0:
            w_row(r+di, '', '46-47', '故障字3', 'unsigned short int', 2, dl, '', hex_of(46, 47), f3_result)
        else:
            w_row(r+di, '', '', '', '', '', dl, '', '', '')
    for col in [2, 3, 4, 5, 8, 9]:
        ws.merge_cells(start_row=f3_start, start_column=col, end_row=f3_start+14, end_column=col)
    r += 15
    
    nav_end = r - 1
    # 合并A列"惯导数据"
    ws.cell(row=nav_start, column=1, value='惯导数据')
    ws.merge_cells(start_row=nav_start, start_column=1, end_row=nav_end, end_column=1)
    
    # ===== 转发卫导数据 =====
    gnss_start = r
    
    def write_gnss_1row(start_r, byte_str, name, dtype, nbytes, desc, note, hx, val_str):
        w_row(start_r, '', byte_str, name, dtype, nbytes, desc, note, hx, val_str)
    
    gnss_note = '卫导无连接/断开默认置0'
    write_gnss_1row(r, '48-51', '水平定位精度（RTK1）', 'int', 4,
                    'D0(LSB)-D15(MSB):\nBNR，无符号\nLSB:0.03125\n其余置零',
                    gnss_note, hex_of(48, 51), f"{parsed['hdop_rtk1']:.5f}"); r+=1
    write_gnss_1row(r, '52-55', '水平定位精度（RTK2）', 'int', 4,
                    'D0(LSB)-D15(MSB):\nBNR，无符号\nLSB:0.03125\n其余置零',
                    gnss_note, hex_of(52, 55), f"{parsed['hdop_rtk2']:.5f}"); r+=1
    write_gnss_1row(r, '56-59', '垂直定位精度（RTK1）', 'int', 4,
                    'D0(LSB)-D15(MSB):\nBNR，无符号\nLSB:0.03125\n其余置零',
                    gnss_note, hex_of(56, 59), f"{parsed['vdop_rtk1']:.5f}"); r+=1
    write_gnss_1row(r, '60-63', '垂直定位精度（RTK2）', 'int', 4,
                    'D0(LSB)-D15(MSB):\nBNR，无符号\nLSB:0.03125\n其余置零',
                    gnss_note, hex_of(60, 63), f"{parsed['vdop_rtk2']:.5f}"); r+=1
    write_gnss_1row(r, '64', '天线定位星数（RTK1）', 'unsigned char', 1,
                    'BNR，无符号\nD0(LSB)-D4（MSB）\n其余位全部置0',
                    gnss_note, f'0x{raw[64]:02X}', str(parsed['sat_rtk1'])); r+=1
    write_gnss_1row(r, '65', '天线定位星数（RTK2）', 'unsigned char', 1,
                    'BNR，无符号\nD0(LSB)-D4（MSB）\n其余位全部置0',
                    gnss_note, f'0x{raw[65]:02X}', str(parsed['sat_rtk2'])); r+=1
    
    diff_desc = 'D0(LSB)-D4（MSB）：\n0x04=无效解\n0x08=单点定位解\n0x0C=伪距差分\n0x15=固定解\n0x0D=浮点解\n其余位全部置0'
    write_gnss_1row(r, '66-67', '差分信息（RTK1）', 'unsigned short int', 2,
                    diff_desc, '卫导无连接/断开默认置：无效解',
                    hex_of(66, 67), parsed['diff_rtk1']); r+=1
    write_gnss_1row(r, '68-69', '差分信息（RTK2）', 'unsigned short int', 2,
                    diff_desc, '卫导无连接/断开默认置：无效解',
                    hex_of(68, 69), parsed['diff_rtk2']); r+=1
    
    valid_desc = '卫导信息状态字D0-D1:\n01:无效; 11:有效\nDOP值信息状态字D2-D3:\n01:无效; 11:有效\n其余位全部置0'
    write_gnss_1row(r, '70', '卫导有效字（RTK1）', 'unsigned char', 1,
                    valid_desc, '卫导无连接/断开默认置：无效',
                    f'0x{raw[70]:02X}',
                    f"卫导：{parsed['gnss_status_rtk1']}\nDOP：{parsed['dop_status_rtk1']}"); r+=1
    write_gnss_1row(r, '71', '卫导有效字（RTK2）', 'unsigned char', 1,
                    valid_desc, '卫导无连接/断开默认置：无效',
                    f'0x{raw[71]:02X}',
                    f"卫导：{parsed['gnss_status_rtk2']}\nDOP：{parsed['dop_status_rtk2']}"); r+=1
    
    gnss_end = r - 1
    ws.cell(row=gnss_start, column=1, value='转发卫导数据')
    ws.merge_cells(start_row=gnss_start, start_column=1, end_row=gnss_end, end_column=1)
    
    # ===== 硬件序列号、软件版本号 =====
    ver_start = r
    sw_desc = ('厂家-SW-Vaa.0bb\naa：软件大版本号\n0bb：软件小版本号\n'
               'D0-D5:小版本号(0-63)\nD6-D9：大版本号(0-15)\nD10-D11：厂家编号')
    w_row(r, '', '72-73', '导航控制软件版本号', 'unsigned short int', 2,
          sw_desc, '', hex_of(72, 73), parsed['sw_version']); r+=1
    
    hw_desc = ('厂家-HW-Vaa.0bb\naa：硬件大版本号\n0bb：硬件小版本号\n'
               'D0-D5:小版本号(0-63)\nD6-D9：大版本号(0-15)\nD10-D11：厂家编号')
    w_row(r, '', '74-75', '硬件版本', 'unsigned short int', 2,
          hw_desc, '', hex_of(74, 75), parsed['hw_version']); r+=1
    
    ver_end = r - 1
    ws.cell(row=ver_start, column=1, value='硬件序列号\n软件版本号')
    ws.merge_cells(start_row=ver_start, start_column=1, end_row=ver_end, end_column=1)
    
    # ===== 预留 + CRC =====
    other_start = r
    w_row(r, '', '76-77', '预留', 'unsigned short int', 2,
          '如不使用置0', '', hex_of(76, 77), str(parsed['reserved'])); r+=1
    
    crc_ok = '校验通过' if parsed['crc_ok'] else '校验失败'
    w_row(r, '', '78-79', 'CRC校验位', 'unsigned char', 2,
          '计算字节：2-77 CRC检验,16位', '',
          hex_of(78, 79),
          f"接收:0x{parsed['crc_recv']:04X}\n计算:0x{parsed['crc_calc']:04X}\n{crc_ok}"); r+=1
    
    other_end = r - 1
    ws.cell(row=other_start, column=1, value='其他')
    ws.merge_cells(start_row=other_start, start_column=1, end_row=other_end, end_column=1)
    
    # ===== 最底部：原始数据完整hex =====
    r += 1
    ws.cell(row=r, column=1, value='原始数据(hex)').font = hdr_font
    hex_full = ' '.join(f'{b:02X}' for b in raw)
    ws.cell(row=r, column=2, value=hex_full).font = data_font
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)


# ==================== 主程序 ====================
def main():
    filepath = find_irs_file()
    if not filepath:
        print("未找到IRS2_1.txt文件！")
        return
    
    print(f"文件: {filepath}")
    print(f"大小: {os.path.getsize(filepath) / 1024 / 1024:.2f} MB")
    
    # 提取前3个包
    print("\n正在提取前3个数据包...")
    packets = extract_packets(filepath, n=3)
    print(f"成功提取 {len(packets)} 个数据包")
    
    for i, pkt in enumerate(packets):
        print(f"\n  包{i+1}: {' '.join(f'{b:02X}' for b in pkt[:10])}... ({len(pkt)} bytes)")
    
    # 解析并生成Excel
    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)
    
    for i, pkt in enumerate(packets):
        print(f"\n正在解析并写入第 {i+1} 个包...")
        parsed = parse_packet(pkt)
        ws = wb.create_sheet(title=f'数据包{i+1}')
        write_sheet(ws, i+1, parsed)
        
        # 打印关键信息
        print(f"  厂家: {parsed['id_mfr']}")
        print(f"  帧计数: {parsed['frame_count']}")
        print(f"  航向: {parsed['heading']:.2f} deg  俯仰: {parsed['pitch']:.2f} deg  滚动: {parsed['roll']:.2f} deg")
        print(f"  纬度: {parsed['latitude']:.7f} deg ({parsed['lat_dir']})  经度: {parsed['longitude']:.7f} deg ({parsed['lon_dir']})")
        print(f"  高度: {parsed['altitude']:.2f} m")
        print(f"  CRC: {'通过' if parsed['crc_ok'] else '失败'}")
    
    # 保存
    output_dir = os.path.dirname(filepath)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'IRS2_1_前3包解析_{timestamp}.xlsx')
    wb.save(output_path)
    print(f"\n完成！文件已保存到：{output_path}")


if __name__ == '__main__':
    main()
