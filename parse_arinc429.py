# -*- coding: utf-8 -*-
"""
ARINC429 解析脚本 - CE-25A 前轮转弯系统通信协议
基于 CE-25A项目0号机前轮转弯系统通信协议EOICD-V5.0 文档

========================================================
ARINC429 数据字格式 (32位整型):
========================================================

  Bit  1-8  : 标签 (Label), 八进制编码 (位反序)
  Bit  9-10 : SDI (源/目标标识符), 部分字中为预留
  Bit 11-29 : 数据域 (具体定义因数据字类型而异)
  Bit 29    : 符号位 (BNR有符号类型: 0=正, 1=负)
  Bit 30-31 : 状态矩阵 (SSM): 00=故障, 01=无效, 10=测试, 11=正常
  Bit 32    : 奇校验位

========================================================
标签(Label)解读规则:
========================================================
  - 8位二进制, ARINC429规定位反序传输
  - 从Bit1到Bit8读取后, 按 前2位|中间3位|后3位 分组 → 八进制3位数
  - 例: bit1~bit8 = 01001101 → 01|001|101 → 1|1|5 → Label 115(八进制)
  - 等价于: 将bits 1-8的字节做位反转, 得到label的十进制值, 再表示为八进制

========================================================
数据域解读规则:
========================================================
  - LSB = 最低有效位, MSB = 最高有效位
  - BNR有符号: 数据位(如bit17-28) + 符号位(bit29) 组成二进制补码
    符号位=0: 正数, 物理值 = 原始值 × 分辨率
    符号位=1: 负数, 补码方式: 数据位取反再加一得到幅值, 物理值取负
  - BNR无符号: 数据位直接为无符号整数, 物理值 = 原始值 × 分辨率
  - 离散量(Discrete): 逐位解读, 每位有独立含义
  - 预留位: 忽略(不参与计算)

========================================================
状态矩阵 (SSM, Bits 30-31):
========================================================
  00 = 故障 (Failure Warning)
  01 = 无效 (No Computed Data)
  10 = 测试 (Functional Test)
  11 = 正常 (Normal Operation)

========================================================
校验位 (Bit 32): 奇校验
========================================================
  整个32位字中'1'的个数应为奇数
"""

import sys
import os
import re
from datetime import datetime

# ============================================================
# 基础工具函数
# ============================================================

def reverse_bits_8(byte_val):
    """反转8位二进制的位序 (ARINC429 Label位反序规则)
    
    ARINC429中Label字段(bit1-bit8)的位序与实际Label值相反:
    bit1是Label最高位, bit8是Label最低位.
    
    Args:
        byte_val: 从32位字中直接提取的bits 1-8 (整数0-255)
    Returns:
        位反转后的Label值 (十进制), 用oct()可转为八进制
    """
    result = 0
    for i in range(8):
        if byte_val & (1 << i):
            result |= (1 << (7 - i))
    return result


def extract_label(word):
    """从32位ARINC429字中提取Label值
    
    Args:
        word: 32位整数 (bit1在最低位, bit32在最高位)
    Returns:
        label_decimal: Label的十进制值
        label_octal_str: Label的八进制字符串 (如 '115')
    """
    raw = word & 0xFF  # 提取bits 1-8
    label_val = reverse_bits_8(raw)
    label_oct_str = oct(label_val)[2:]  # 去掉 '0o' 前缀
    return label_val, label_oct_str


def extract_bit(word, bit_num):
    """提取ARINC429字中指定位的值
    
    Args:
        word: 32位整数
        bit_num: ARINC429位号 (1-32, bit1在最低位)
    Returns:
        0 或 1
    """
    return (word >> (bit_num - 1)) & 1


def extract_bits(word, start_bit, end_bit):
    """提取ARINC429字中指定范围的位值 (含两端)
    
    Args:
        word: 32位整数
        start_bit: 起始位号 (1-based, 含)
        end_bit: 结束位号 (1-based, 含)
    Returns:
        提取的无符号整数值 (start_bit为LSB)
    """
    num_bits = end_bit - start_bit + 1
    mask = (1 << num_bits) - 1
    return (word >> (start_bit - 1)) & mask


def check_odd_parity(word):
    """检查32位字的奇校验
    
    Returns:
        True: 校验通过 (1的个数为奇数)
        False: 校验失败
    """
    count = bin(word & 0xFFFFFFFF).count('1')
    return count % 2 == 1


def decode_ssm(ssm_val):
    """解码状态矩阵 (SSM, Bits 30-31)
    
    按协议文档定义:
      00 → 故障
      01 → 无效
      10 → 测试
      11 → 正常
    
    Args:
        ssm_val: 0-3 (bit31为高位, bit30为低位)
    Returns:
        状态描述字符串, 格式: "XX-含义"
    """
    ssm_map = {
        0b00: '00-故障',
        0b01: '01-无效',
        0b10: '10-测试',
        0b11: '11-正常'
    }
    return ssm_map.get(ssm_val, f'{ssm_val:02b}-未知')


def decode_bnr_signed(word, data_start, data_end, sign_bit, resolution):
    """解码BNR有符号数据
    
    使用二进制补码:
    - 符号位=0: 正数, 值 = data × resolution
    - 符号位=1: 负数, 将(符号位+数据位)视为补码整数
    
    Args:
        word: 32位ARINC429字
        data_start: 数据域起始位 (LSB)
        data_end: 数据域结束位 (MSB)
        sign_bit: 符号位位号
        resolution: 分辨率 (物理量/bit)
    Returns:
        (raw_data, sign, physical_value)
    """
    data_raw = extract_bits(word, data_start, data_end)
    sign = extract_bit(word, sign_bit)
    num_data_bits = data_end - data_start + 1
    
    # 组合符号位和数据位, 形成补码整数
    combined = (sign << num_data_bits) | data_raw
    total_bits = num_data_bits + 1
    
    if sign:
        # 负数: 二进制补码转换
        signed_val = combined - (1 << total_bits)
    else:
        signed_val = combined
    
    physical_val = signed_val * resolution
    return data_raw, sign, physical_val


def decode_bnr_unsigned(word, data_start, data_end, resolution):
    """解码BNR无符号数据
    
    Args:
        word: 32位ARINC429字
        data_start: 数据域起始位 (LSB)
        data_end: 数据域结束位 (MSB)
        resolution: 分辨率 (物理量/bit)
    Returns:
        (raw_data, physical_value)
    """
    data_raw = extract_bits(word, data_start, data_end)
    physical_val = data_raw * resolution
    return data_raw, physical_val


def interpret_discrete_desc(desc_str, bit_val):
    """根据离散位描述字符串和当前位值, 返回字段名和含义解释
    
    描述格式示例: "调零指令: 0=调零, 1=无效"
                 "转弯断开开关: 1=转弯断开, 0=无效 (驾驶舱显控开关)"
    
    Args:
        desc_str: 协议中对该位的描述 (如 "调零指令: 0=调零, 1=无效")
        bit_val: 当前位的值 (0 或 1)
    Returns:
        (field_name, interpretation): 字段名称和当前值的含义
    """
    # 提取字段名 (冒号前面的部分)
    if ':' in desc_str:
        field_name = desc_str.split(':')[0].strip()
        rest = desc_str.split(':', 1)[1].strip()
    elif '：' in desc_str:
        field_name = desc_str.split('：')[0].strip()
        rest = desc_str.split('：', 1)[1].strip()
    else:
        field_name = desc_str
        rest = ''
    
    # 尝试从描述中解析 "0=xxx, 1=yyy" 格式
    interpretation = ''
    # 匹配 "0=xxx" 和 "1=yyy" 模式
    matches = re.findall(r'(\d)\s*[=＝]\s*([^,，;；()（）]+)', rest)
    val_map = {}
    for m_val, m_desc in matches:
        val_map[int(m_val)] = m_desc.strip()
    
    if bit_val in val_map:
        interpretation = val_map[bit_val]
    else:
        interpretation = str(bit_val)
    
    return field_name, interpretation


# ============================================================
# 协议定义 - 所有数据字类型
# ============================================================

# --- RDIU → SCU (飞控发送的) ---

LABEL_115 = {
    'label_oct': '115',
    'label_dec': 0o115,  # = 77
    'name': '脚蹬转弯指令',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 1'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',  # 度
    'range': '\u00b160\u00b0',
    'resolution': 0.014653,
    'encoding': 'BNR',
    'data_bits': (17, 28),  # 12 bits, bit17=LSB, bit28=MSB
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': '数据字1, 来自FCM 1'
}

LABEL_116 = {
    'label_oct': '116',
    'label_dec': 0o116,  # = 78
    'name': '脚蹬转弯指令',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 2'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',
    'range': '\u00b160\u00b0',
    'resolution': 0.014653,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': '数据字2, 来自FCM 2'
}

LABEL_117 = {
    'label_oct': '117',
    'label_dec': 0o117,  # = 79
    'name': '脚蹬转弯指令',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 3'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',
    'range': '\u00b160\u00b0',
    'resolution': 0.014653,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': '数据字3, 来自FCM 3'
}

LABEL_120 = {
    'label_oct': '120',
    'label_dec': 0o120,  # = 80 (117后的下一个有效八进制数)
    'name': '脚蹬转弯指令',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 4'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',
    'range': '\u00b160\u00b0',
    'resolution': 0.014653,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': '数据字4, 来自FCM 4 (八进制117之后是120, 不存在118/119)'
}

LABEL_270 = {
    'label_oct': '270',
    'label_dec': 0o270,  # = 184
    'name': '驾驶舱信号（驾驶舱转弯断开开关）',
    'direction': 'RDIU -> SCU',
    'sources': ['驾驶舱'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        15: '转弯断开开关: 1=转弯断开, 0=无效 (驾驶舱显控开关)',
    },
    'reserved_bits': '9-14, 16-29',
    'notes': '数据字5, 来自驾驶舱'
}

LABEL_354 = {
    'label_oct': '354',
    'label_dec': 0o354,  # = 236
    'name': '控制信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 1'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        14: '调零指令: 0=调零, 1=无效',
        15: '转弯断开开关: 1=转弯断开, 0=无效 (方舱开关)',
    },
    'reserved_bits': '9-13, 16-29',
    'notes': '数据字6, 来自FCM 1'
}

LABEL_355 = {
    'label_oct': '355',
    'label_dec': 0o355,  # = 237
    'name': '控制信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 2'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        14: '调零指令: 0=调零, 1=无效',
        15: '转弯断开开关: 1=转弯断开, 0=无效 (方舱开关)',
    },
    'reserved_bits': '9-13, 16-29',
    'notes': '数据字7, 来自FCM 2'
}

LABEL_356 = {
    'label_oct': '356',
    'label_dec': 0o356,  # = 238
    'name': '控制信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 3'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        14: '调零指令: 0=调零, 1=无效',
        15: '转弯断开开关: 1=转弯断开, 0=无效 (方舱开关)',
    },
    'reserved_bits': '9-13, 16-29',
    'notes': '数据字8, 来自FCM 3'
}

LABEL_357 = {
    'label_oct': '357',
    'label_dec': 0o357,  # = 239
    'name': '控制信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 4'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        14: '调零指令: 0=调零, 1=无效',
        15: '转弯断开开关: 1=转弯断开, 0=无效 (方舱开关)',
    },
    'reserved_bits': '9-13, 16-29',
    'notes': '数据字9, 来自FCM 4'
}

# 空速信号 - 4个来源共享相同结构, 不同Label
LABEL_374 = {
    'label_oct': '374',
    'label_dec': 0o374,  # = 252
    'name': '空速信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 1'],
    'data_type': 'BNR_UNSIGNED',
    'unit': 'Knot',
    'range': '0-270 Knot',
    'resolution': 0.125,
    'encoding': 'BNR',
    'data_bits': (15, 28),  # 14 bits, bit15=LSB(0.125), bit28=MSB(1024)
    'sign_bit': None,
    'reserved_bits': '9-14',
    'notes': '数据字10, 来自FCM 1'
}

LABEL_375 = {
    'label_oct': '375',
    'label_dec': 0o375,  # = 253
    'name': '空速信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 2'],
    'data_type': 'BNR_UNSIGNED',
    'unit': 'Knot',
    'range': '0-270 Knot',
    'resolution': 0.125,
    'encoding': 'BNR',
    'data_bits': (15, 28),
    'sign_bit': None,
    'reserved_bits': '9-14',
    'notes': '数据字11, 来自FCM 2'
}

LABEL_376 = {
    'label_oct': '376',
    'label_dec': 0o376,  # = 254
    'name': '空速信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 3'],
    'data_type': 'BNR_UNSIGNED',
    'unit': 'Knot',
    'range': '0-270 Knot',
    'resolution': 0.125,
    'encoding': 'BNR',
    'data_bits': (15, 28),
    'sign_bit': None,
    'reserved_bits': '9-14',
    'notes': '数据字12, 来自FCM 3'
}

LABEL_377 = {
    'label_oct': '377',
    'label_dec': 0o377,  # = 255
    'name': '空速信号',
    'direction': 'RDIU -> SCU',
    'sources': ['FCM 4'],
    'data_type': 'BNR_UNSIGNED',
    'unit': 'Knot',
    'range': '0-270 Knot',
    'resolution': 0.125,
    'encoding': 'BNR',
    'data_bits': (15, 28),
    'sign_bit': None,
    'reserved_bits': '9-14',
    'notes': '数据字13, 来自FCM 4'
}

# --- SCU → RDIU (飞控接收的) ---

LABEL_111 = {
    'label_oct': '111',
    'label_dec': 0o111,  # = 73
    'name': '前轮角度反馈',
    'direction': 'SCU -> RDIU',
    'sources': ['SCU'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',
    'range': '\u00b160\u00b0',
    'resolution': 0.014653,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': 'SCU数据字1'
}

LABEL_112 = {
    'label_oct': '112',
    'label_dec': 0o112,  # = 74
    'name': '左手轮指令',
    'direction': 'SCU -> RDIU',
    'sources': ['SCU'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',
    'range': '\u00b190\u00b0',
    'resolution': 0.021978,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': 'SCU数据字2'
}

LABEL_113 = {
    'label_oct': '113',
    'label_dec': 0o113,  # = 75
    'name': '右手轮指令',
    'direction': 'SCU -> RDIU',
    'sources': ['SCU'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',
    'range': '\u00b190\u00b0',
    'resolution': 0.021978,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': 'SCU数据字3'
}

LABEL_114 = {
    'label_oct': '114',
    'label_dec': 0o114,  # = 76
    'name': '零位历史值（累加值）',
    'direction': 'SCU -> RDIU',
    'sources': ['SCU'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',
    'range': '\u00b145\u00b0',
    'resolution': 0.010986,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': 'SCU数据字4'
}

LABEL_244 = {
    'label_oct': '244',
    'label_dec': 0o244,  # = 164
    'name': '状态上报',
    'direction': 'SCU -> RDIU',
    'sources': ['SCU'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        # bits 17-18: 牵引状态 已在special_fields中定义
        12: '脚蹬解除状态: 1=解除状态, 0=正常状态',
    },
    'special_fields': [
        {'name': '前轮工作状态上报', 'bits': (14, 16), 'values': {
            0b010: '转弯状态',
            0b011: '减摆状态',
            0b100: '调零状态',
        }},
        {'name': '牵引状态', 'bits': (17, 18), 'values': {
            0b00: '牵引无效状态',
            0b01: '允许牵引状态',
            0b10: '禁止牵引状态',
            0b11: '(未定义)',
        }},
        {'name': '软件大版本', 'bits': (19, 22), 'type': 'uint'},
        {'name': '软件小版本', 'bits': (23, 26), 'type': 'uint'},
    ],
    'reserved_bits': '9-11, 13, 27-29',
    'notes': 'SCU数据字5, 前轮工作状态: 010=转弯, 011=减摆, 100=调零'
}

LABEL_314 = {
    'label_oct': '314',
    'label_dec': 0o314,  # = 204
    'name': '故障状态字',
    'direction': 'SCU -> RDIU',
    'sources': ['SCU'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        14: '左手轮传感器和值故障: 1=故障, 0=正常',
        15: '右手轮传感器和值故障: 1=故障, 0=正常',
        16: 'ARINC429通讯故障: 1=故障, 0=正常',
        17: '前轮工作故障: 1=故障, 0=正常',
        18: '牵引超行程: 1=故障, 0=正常',
    },
    'reserved_bits': '9-13, 19-29',
    'notes': 'SCU数据字6, 故障字"1"为故障"0"为正常, 牵引超行程:舵机前轮位置反馈超过+/-50度后上报'
}

LABEL_212 = {
    'label_oct': '212',
    'label_dec': 0o212,  # = 138
    'name': '脚蹬转弯指令回绕',
    'direction': 'SCU -> RDIU',
    'sources': ['SCU'],
    'data_type': 'BNR_SIGNED',
    'unit': '\u00b0',  # 度 (回绕Label 115, 单位与脚蹬转弯指令一致)
    'range': '\u00b160\u00b0',
    'resolution': 0.014653,
    'encoding': 'BNR',
    'data_bits': (17, 28),
    'sign_bit': 29,
    'reserved_bits': '9-16',
    'notes': 'SCU回报RDIU的ARINC429数据字1'
}

LABEL_154 = {
    'label_oct': '154',
    'label_dec': 0o154,  # = 108
    'name': '控制信号',
    'direction': 'SCU -> RDIU',
    'sources': ['RDIU'],
    'data_type': 'DISCRETE',
    'unit': '',
    'range': '',
    'resolution': None,
    'encoding': 'BNR',
    'discrete_bits': {
        13: '调零指令: 1=调零, 0=无效',
        14: '转弯断开开关: 1=转弯断开, 0=无效',
        15: 'BCMU_PARK_BRK_ON: 1=有效, 0=无效',
    },
    'reserved_bits': '9-12, 16-29',
    'notes': 'SCU回报RDIU的ARINC429数据字4'
}

# ============================================================
# Label查找表 (按label十进制值索引)
# ============================================================

LABEL_LOOKUP = {}
for _def in [LABEL_115, LABEL_116, LABEL_117, LABEL_120,
             LABEL_270,
             LABEL_354, LABEL_355, LABEL_356, LABEL_357,
             LABEL_374, LABEL_375, LABEL_376, LABEL_377,
             LABEL_111, LABEL_112, LABEL_113, LABEL_114,
             LABEL_244, LABEL_314, LABEL_212, LABEL_154]:
    LABEL_LOOKUP[_def['label_dec']] = _def


# ============================================================
# 核心解析函数
# ============================================================

def parse_arinc429_word(word):
    """完整解析一个32位ARINC429数据字
    
    Args:
        word: 32位整数 (bit1在最低位)
    Returns:
        dict: 解析结果, 包含所有字段信息
    """
    result = {}
    
    # 1. 原始数据
    result['raw_hex'] = f'0x{word:08X}'
    # 按协议传输顺序显示: bit01(最先发送)在左, bit32在右
    # 标准二进制是 bit32在左(MSB), 需要翻转
    result['raw_bin'] = f'{word:032b}'[::-1]
    
    # 2. 提取Label
    label_dec, label_oct = extract_label(word)
    result['label_dec'] = label_dec
    result['label_oct'] = label_oct
    
    # 3. 提取SDI (bits 9-10)
    sdi = extract_bits(word, 9, 10)
    result['sdi'] = sdi
    
    # 4. 提取SSM (bits 30-31)
    ssm = extract_bits(word, 30, 31)
    result['ssm_raw'] = ssm
    result['ssm_desc'] = decode_ssm(ssm)
    
    # 5. 提取奇校验位 (bit 32)
    parity_bit = extract_bit(word, 32)
    result['parity_bit'] = parity_bit
    result['parity_ok'] = check_odd_parity(word)
    
    # 6. 查找协议定义
    word_def = LABEL_LOOKUP.get(label_dec)
    if word_def:
        result['known'] = True
        result['name'] = word_def['name']
        result['direction'] = word_def['direction']
        result['sources'] = word_def['sources']
        result['unit'] = word_def.get('unit', '')
        result['range'] = word_def.get('range', '')
        result['notes'] = word_def.get('notes', '')
        
        # 7. 根据数据类型解码
        dtype = word_def['data_type']
        
        if dtype == 'BNR_SIGNED':
            ds, de = word_def['data_bits']
            sb = word_def['sign_bit']
            res = word_def['resolution']
            data_raw, sign, phys_val = decode_bnr_signed(word, ds, de, sb, res)
            result['data_raw'] = data_raw
            result['sign'] = sign
            result['sign_desc'] = '正' if sign == 0 else '负'
            result['physical_value'] = phys_val
            result['physical_str'] = f'{phys_val:.6f} {word_def["unit"]}'
            result['resolution'] = res
            result['data_bits_range'] = f'bit{ds}-bit{de} ({de-ds+1}位)'
            
        elif dtype == 'BNR_UNSIGNED':
            ds, de = word_def['data_bits']
            res = word_def['resolution']
            data_raw, phys_val = decode_bnr_unsigned(word, ds, de, res)
            result['data_raw'] = data_raw
            result['sign'] = None
            result['physical_value'] = phys_val
            result['physical_str'] = f'{phys_val:.4f} {word_def["unit"]}'
            result['resolution'] = res
            result['data_bits_range'] = f'bit{ds}-bit{de} ({de-ds+1}位)'
            
        elif dtype == 'DISCRETE':
            # 解码各离散位
            discrete_results = []
            if 'discrete_bits' in word_def:
                for bit_num, desc in sorted(word_def['discrete_bits'].items()):
                    bit_val = extract_bit(word, bit_num)
                    discrete_results.append({
                        'bit': bit_num,
                        'value': bit_val,
                        'description': desc
                    })
            
            # 解码特殊多位字段
            special_results = []
            if 'special_fields' in word_def:
                for sf in word_def['special_fields']:
                    bs, be = sf['bits']
                    field_val = extract_bits(word, bs, be)
                    if 'values' in sf:
                        val_desc = sf['values'].get(field_val, f'未定义({field_val})')
                        special_results.append({
                            'name': sf['name'],
                            'bits': f'bit{bs}-bit{be}',
                            'raw_value': field_val,
                            'description': val_desc
                        })
                    elif sf.get('type') == 'uint':
                        special_results.append({
                            'name': sf['name'],
                            'bits': f'bit{bs}-bit{be}',
                            'raw_value': field_val,
                            'description': str(field_val)
                        })
            
            result['discrete_bits'] = discrete_results
            result['special_fields'] = special_results
    else:
        result['known'] = False
        result['name'] = f'未知Label ({label_oct} oct)'
        # 输出所有位的原始值供参考
        all_bits = {}
        for i in range(1, 33):
            all_bits[i] = extract_bit(word, i)
        result['all_bits'] = all_bits
    
    return result


def format_parse_result(result):
    """将解析结果格式化为可读字符串
    
    Args:
        result: parse_arinc429_word()返回的dict
    Returns:
        格式化的多行字符串
    """
    lines = []
    lines.append('=' * 60)
    lines.append(f'ARINC429 数据字解析结果')
    lines.append('=' * 60)
    lines.append(f'原始数据 (HEX): {result["raw_hex"]}')
    lines.append(f'原始数据 (BIN): {result["raw_bin"]}')
    lines.append(f'')
    lines.append(f'--- 标签 (Label) ---')
    lines.append(f'  Label (八进制): {result["label_oct"]}')
    lines.append(f'  Label (十进制): {result["label_dec"]}')
    lines.append(f'')
    lines.append(f'--- SDI (Bits 9-10) ---')
    lines.append(f'  SDI: {result["sdi"]:02b} ({result["sdi"]})')
    lines.append(f'')
    lines.append(f'--- 状态矩阵 SSM (Bits 30-31) ---')
    lines.append(f'  SSM: {result["ssm_raw"]:02b} -> {result["ssm_desc"]}')
    lines.append(f'')
    lines.append(f'--- 奇校验 (Bit 32) ---')
    lines.append(f'  校验位: {result["parity_bit"]}')
    lines.append(f'  校验结果: {"通过" if result["parity_ok"] else "失败"}')
    lines.append(f'')
    
    if result['known']:
        lines.append(f'--- 信号识别 ---')
        lines.append(f'  信号名称: {result["name"]}')
        lines.append(f'  方向: {result["direction"]}')
        lines.append(f'  可能来源: {", ".join(result["sources"])}')
        if result.get('range'):
            lines.append(f'  信号范围: {result["range"]}')
        if result.get('notes'):
            lines.append(f'  备注: {result["notes"]}')
        lines.append(f'')
        
        if 'physical_value' in result:
            lines.append(f'--- 数据解析 ---')
            lines.append(f'  数据位范围: {result.get("data_bits_range", "N/A")}')
            lines.append(f'  原始数据值: {result["data_raw"]} (0x{result["data_raw"]:X})')
            if result.get('sign') is not None:
                lines.append(f'  符号位: {result["sign"]} ({result["sign_desc"]})')
            lines.append(f'  分辨率: {result["resolution"]}')
            lines.append(f'  物理值: {result["physical_str"]}')
        
        if 'discrete_bits' in result and result['discrete_bits']:
            lines.append(f'--- 离散位解析 ---')
            for db in result['discrete_bits']:
                field_name, interp = interpret_discrete_desc(db['description'], db['value'])
                lines.append(f'  Bit {db["bit"]}: {db["value"]} -> [{field_name}] 当前值含义: {interp}  (定义: {db["description"]})')
        
        if 'special_fields' in result and result['special_fields']:
            lines.append(f'--- 特殊字段解析 ---')
            for sf in result['special_fields']:
                lines.append(f'  {sf["name"]} ({sf["bits"]}): {sf["raw_value"]} -> {sf["description"]}')
    else:
        lines.append(f'--- 未识别的Label ---')
        lines.append(f'  此Label未在CE-25A前轮转弯系统协议中定义')
        if 'all_bits' in result:
            bit_str = ''.join(str(result['all_bits'][i]) for i in range(32, 0, -1))
            lines.append(f'  全部位 (bit32→bit1): {bit_str}')
    
    lines.append('=' * 60)
    return '\n'.join(lines)


def parse_hex_input(hex_str):
    """解析用户输入的十六进制字符串
    
    支持格式:
    - "67FF00B2"       (纯hex, 8位)
    - "0x67FF00B2"     (带0x前缀)
    - "B2 00 FF 67"    (4个空格分隔的字节, 小端序: byte0=bit1-8在前)
    
    Args:
        hex_str: 十六进制字符串
    Returns:
        32位整数
    """
    hex_str = hex_str.strip()
    if hex_str.startswith(('0x', '0X')):
        hex_str = hex_str[2:]
    
    # 检查是否是空格分隔的字节格式 (如 "B2 00 FF 67")
    parts = hex_str.split()
    if len(parts) == 4 and all(len(p) == 2 for p in parts):
        # 4个字节, 小端序: byte0=bits1-8(Label), byte3=bits25-32(SSM+Parity)
        b = bytes([int(p, 16) for p in parts])
        return int.from_bytes(b, byteorder='little')
    
    return int(hex_str, 16)


def parse_binary_input(bin_str):
    """解析用户输入的二进制字符串
    
    支持格式:
    - "01001101..." (32位二进制)
    - "0b01001101..." (带0b前缀)
    
    注意: 输入采用标准数学顺序, 即 bit32(MSB)在左, bit1(LSB)在右.
          这与Excel中 raw_bin 列的显示顺序(bit1在左)相反!
          若要从Excel中复制二进制字符串使用, 需先反转.
    
    Args:
        bin_str: 二进制字符串
    Returns:
        32位整数
    """
    bin_str = bin_str.strip()
    if bin_str.startswith(('0b', '0B')):
        bin_str = bin_str[2:]
    return int(bin_str, 2)


def load_raw_byte_file(filepath):
    """从原始字节文件中读取ARINC429数据字
    
    支持格式: 空格分隔的十六进制字节流, 每4个字节为一个32位ARINC429字
    例: "B2 00 FF 67 92 00 FE 6B ..."
    
    字节序: 小端 (byte0=bits1-8 即Label, byte3=bits25-32 即SSM+Parity)
    
    Args:
        filepath: 原始数据文件路径
    Returns:
        32位整数列表
    """
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    
    # 提取所有合法的十六进制token
    tokens = []
    for token in content.split():
        token = token.strip()
        if len(token) == 2:
            try:
                int(token, 16)
                tokens.append(token)
            except ValueError:
                continue
    
    # 每4个字节组成一个32位ARINC429字 (小端序)
    words = []
    for i in range(0, len(tokens) - 3, 4):
        try:
            b = bytes([int(tokens[i], 16), int(tokens[i+1], 16),
                        int(tokens[i+2], 16), int(tokens[i+3], 16)])
            word = int.from_bytes(b, byteorder='little')
            words.append(word)
        except (ValueError, IndexError):
            continue
    
    return words


# ============================================================
# 批量解析 & Excel输出
# ============================================================

def parse_batch_to_excel(words, output_path=None):
    """批量解析ARINC429字并输出到Excel
    
    Args:
        words: 32位整数列表
        output_path: 输出Excel路径 (None则自动生成)
    Returns:
        输出文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("需要安装openpyxl: pip install openpyxl")
        return None
    
    if output_path is None:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_path = os.path.join(script_dir, f'ARINC429_解析结果_{ts}.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = '解析结果'
    
    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写表头
    headers = [
        '序号', '原始数据(HEX)', '原始4字节(小端)', '32位二进制',
        'Label(八进制)', '信号名称', '方向',
        '数据类型', '原始数据值', '符号', '物理值', '单位',
        'SSM状态', '奇校验', '备注'
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 解析每个字并写入
    for idx, word in enumerate(words, 1):
        result = parse_arinc429_word(word)
        row = idx + 1
        col = 1
        
        # 序号
        ws.cell(row=row, column=col, value=idx).border = thin_border; col += 1
        # 原始HEX (32位整数)
        ws.cell(row=row, column=col, value=result['raw_hex']).border = thin_border; col += 1
        # 原始4字节(小端序): byte0 byte1 byte2 byte3
        raw_bytes = (word & 0xFFFFFFFF).to_bytes(4, byteorder='little')
        raw_4byte_str = ' '.join(f'{b:02X}' for b in raw_bytes)
        ws.cell(row=row, column=col, value=raw_4byte_str).border = thin_border; col += 1
        # 32位二进制
        ws.cell(row=row, column=col, value=result['raw_bin']).border = thin_border; col += 1
        # Label
        ws.cell(row=row, column=col, value=result['label_oct']).border = thin_border; col += 1
        # 信号名称
        ws.cell(row=row, column=col, value=result['name']).border = thin_border; col += 1
        # 方向
        ws.cell(row=row, column=col, value=result.get('direction', '')).border = thin_border; col += 1
        
        # 数据类型 & 值 (columns 8-12)
        if result['known']:
            word_def = LABEL_LOOKUP[result['label_dec']]
            ws.cell(row=row, column=col, value=word_def['data_type']).border = thin_border; col += 1
            
            if 'physical_value' in result:
                ws.cell(row=row, column=col, value=result.get('data_raw', '')).border = thin_border; col += 1
                sign_str = result.get('sign_desc', '')
                ws.cell(row=row, column=col, value=sign_str).border = thin_border; col += 1
                ws.cell(row=row, column=col, value=round(result['physical_value'], 6)).border = thin_border; col += 1
                ws.cell(row=row, column=col, value=result.get('unit', '')).border = thin_border; col += 1
            elif 'discrete_bits' in result:
                # 离散量: 原始数据值 - 显示各位的定义和值
                raw_parts = []
                interp_parts = []
                for db in result['discrete_bits']:
                    field_name, interp = interpret_discrete_desc(db['description'], db['value'])
                    raw_parts.append(f'Bit{db["bit"]}={db["value"]}({field_name})')
                    interp_parts.append(f'{field_name}: {interp}')
                if 'special_fields' in result:
                    for sf in result['special_fields']:
                        raw_val = sf['raw_value']
                        # 对多位字段显示二进制值
                        raw_parts.append(f'{sf["name"]}({sf["bits"]})={raw_val}')
                        interp_parts.append(f'{sf["name"]}: {sf["description"]}')
                ws.cell(row=row, column=col, value='; '.join(raw_parts)).border = thin_border; col += 1
                # 符号列: 空
                ws.cell(row=row, column=col, value='').border = thin_border; col += 1
                # 物理值列: 各字段的含义解释
                ws.cell(row=row, column=col, value='; '.join(interp_parts)).border = thin_border; col += 1
                ws.cell(row=row, column=col, value='').border = thin_border; col += 1
        else:
            ws.cell(row=row, column=col, value='未知').border = thin_border; col += 1
            for _ in range(4):
                ws.cell(row=row, column=col, value='').border = thin_border; col += 1
        
        # SSM状态
        ws.cell(row=row, column=col, value=result['ssm_desc']).border = thin_border; col += 1
        # 校验
        parity_str = '通过' if result['parity_ok'] else '失败'
        cell = ws.cell(row=row, column=col, value=parity_str)
        cell.border = thin_border
        if not result['parity_ok']:
            cell.font = Font(color='FF0000', bold=True)
        col += 1
        # 备注
        ws.cell(row=row, column=col, value=result.get('notes', '')).border = thin_border
    
    # 调整列宽
    from openpyxl.utils import get_column_letter
    col_widths = [6, 14, 16, 38, 12, 25, 14, 14, 45, 6, 45, 8, 25, 8, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    wb.save(output_path)
    return output_path


# ============================================================
# 交互式主程序
# ============================================================

def print_all_labels():
    """打印所有已定义的Label列表"""
    print('\n已定义的ARINC429 Label列表:')
    print('-' * 70)
    print(f'{"Label(Oct)":<12} {"名称":<25} {"方向":<15} {"数据类型":<15}')
    print('-' * 70)
    
    sorted_defs = sorted(LABEL_LOOKUP.values(), key=lambda x: x['label_dec'])
    for d in sorted_defs:
        print(f'{d["label_oct"]:<12} {d["name"]:<25} {d["direction"]:<15} {d["data_type"]:<15}')
    print('-' * 70)


def interactive_mode():
    """交互式解析模式"""
    print('=' * 60)
    print('ARINC429 数据字解析器')
    print('CE-25A 前轮转弯系统通信协议 EOICD-V5.0')
    print('=' * 60)
    print()
    print('使用方法:')
    print('  输入32位HEX: 67FF00B2 或 0x67FF00B2')
    print('  输入4字节:   B2 00 FF 67 (小端序, Label字节在前)')
    print('  输入 "list" 查看所有已定义的Label')
    print('  输入 "quit" 或 "exit" 退出')
    print('  输入 "file <路径>" 从文件批量解析 (每行一个hex值)')
    print('  输入 "raw <路径>"  从原始字节文件解析 (连续字节流)')
    print()
    
    while True:
        try:
            user_input = input('请输入ARINC429数据字 (hex): ').strip()
        except (EOFError, KeyboardInterrupt):
            print('\n退出.')
            break
        
        if not user_input:
            continue
        
        if user_input.lower() in ('quit', 'exit', 'q'):
            print('退出.')
            break
        
        if user_input.lower() == 'list':
            print_all_labels()
            continue
        
        if user_input.lower().startswith('raw '):
            filepath = user_input[4:].strip()
            if os.path.exists(filepath):
                print(f'正在从原始字节文件读取: {filepath}')
                words = load_raw_byte_file(filepath)
                if words:
                    print(f'读取到 {len(words)} 个数据字 (每4字节一个), 正在解析...')
                    for i, w in enumerate(words, 1):
                        raw_bytes = (w & 0xFFFFFFFF).to_bytes(4, byteorder='little')
                        byte_str = ' '.join(f'{b:02X}' for b in raw_bytes)
                        print(f'\n--- 第{i}个字 [{byte_str}] ---')
                        result = parse_arinc429_word(w)
                        print(format_parse_result(result))
                    out = parse_batch_to_excel(words)
                    if out:
                        print(f'\nExcel已保存到: {out}')
                else:
                    print('文件中没有有效数据')
            else:
                print(f'文件不存在: {filepath}')
            continue
        
        if user_input.lower().startswith('file '):
            filepath = user_input[5:].strip()
            if os.path.exists(filepath):
                print(f'正在从文件读取: {filepath}')
                words = []
                with open(filepath, 'r') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            try:
                                words.append(parse_hex_input(line))
                            except ValueError:
                                print(f'  跳过无效行: {line}')
                if words:
                    print(f'读取到 {len(words)} 个数据字, 正在解析...')
                    out = parse_batch_to_excel(words)
                    if out:
                        print(f'结果已保存到: {out}')
                else:
                    print('文件中没有有效数据')
            else:
                print(f'文件不存在: {filepath}')
            continue
        
        # 尝试解析hex输入
        try:
            word = parse_hex_input(user_input)
            result = parse_arinc429_word(word)
            print(format_parse_result(result))
        except ValueError as e:
            print(f'输入格式错误: {e}')
            print('请输入有效的32位十六进制数 (如: 67FF00B2 或 B2 00 FF 67)')


# ============================================================
# 命令行入口
# ============================================================

if __name__ == '__main__':
    if len(sys.argv) > 1:
        # 命令行参数模式
        arg = sys.argv[1]
        
        if arg == '--list':
            print_all_labels()
        elif arg == '--raw' and len(sys.argv) > 2:
            # 原始字节文件模式: 空格分隔的hex字节流, 每4字节=1个ARINC429字
            filepath = sys.argv[2]
            if os.path.exists(filepath):
                words = load_raw_byte_file(filepath)
                if words:
                    print(f'从原始文件读取到 {len(words)} 个ARINC429数据字')
                    out = parse_batch_to_excel(words)
                    if out:
                        print(f'结果已保存到: {out}')
            else:
                print(f'文件不存在: {filepath}')
        elif arg == '--file' and len(sys.argv) > 2:
            filepath = sys.argv[2]
            if os.path.exists(filepath):
                words = []
                with open(filepath, 'r') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            try:
                                words.append(parse_hex_input(line))
                            except ValueError:
                                pass
                if words:
                    out = parse_batch_to_excel(words)
                    if out:
                        print(f'结果已保存到: {out}')
            else:
                print(f'文件不存在: {filepath}')
        elif arg == '--help':
            print('用法:')
            print('  python parse_arinc429.py              # 交互模式')
            print('  python parse_arinc429.py <hex>         # 解析单个数据字 (如 67FF00B2)')
            print('  python parse_arinc429.py "B2 00 FF 67" # 解析4字节(小端序)')
            print('  python parse_arinc429.py --list        # 列出所有Label')
            print('  python parse_arinc429.py --file <f>    # 从文件批量解析(每行一个hex)')
            print('  python parse_arinc429.py --raw <f>     # 从原始字节文件解析(连续字节流)')
        else:
            # 直接解析单个hex值 (支持 "67FF00B2" 或 "B2 00 FF 67")
            try:
                # 如果多个参数, 尝试拼接为空格分隔的4字节
                if len(sys.argv) == 5:
                    combined = ' '.join(sys.argv[1:5])
                    word = parse_hex_input(combined)
                else:
                    word = parse_hex_input(arg)
                result = parse_arinc429_word(word)
                print(format_parse_result(result))
            except ValueError as e:
                print(f'输入格式错误: {e}')
    else:
        # 交互模式
        interactive_mode()
